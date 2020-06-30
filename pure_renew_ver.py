from openpyxl import load_workbook
from openpyxl import Workbook
from konlpy.tag import Kkma
from konlpy.utils import pprint
import gensim

kkma = Kkma()
load_wb = load_workbook("C:\\Users\kimsungju\Desktop\python file\Ahn_0403_0407.xlsx", data_only=True)
load_ws = load_wb['sheet']
wb = Workbook()
sheet1 = wb.active
file_name = 'Ahn0403.xlsx'
sheet1.title = 'sampleSheet'
for i in range(3209):
    num = 'Q' + str(i+1)
    col = load_ws[num].value
    preprocessing = gensim.utils.simple_preprocess(col)
    process_list = []
    for word in preprocessing:
        process = kkma.pos(word)
        exception = ['문재인','더불어민주당', '국민의당','새누리당','안철수','홍준표','김종인','추미애','지지율','바른미래당','심상정','안희정']
        if word in exception :
            process_list.append(word)
        else :
           for j,k in process :
                if k == 'VV' or 'VA' :
                    process_list.append(j)
                    pass 
                elif k == ('JKO' or 'JX' or 'JKM' or 'JKG' or 'JKS') :
                    word = word.replace(j,'')
                    process_list.append(word)
    final = ''
    for fin in process_list:
        final = final + ' ' + fin
    sheet1.cell(row=i+1, column=1).value = i+1
    sheet1.cell(row=i+1, column=2).value = final
wb.save(filename=file_name)
print('Done')




