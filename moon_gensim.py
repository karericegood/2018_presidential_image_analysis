from gensim.models import Word2Vec
import gensim.utils
from openpyxl import load_workbook
from openpyxl import Workbook
from sklearn.manifold import TSNE
import matplotlib as mpl
import matplotlib.pyplot as plt
import gensim
import gensim.models as g
import pandas as pd

# load previous file.
text = []
load_wb1 = load_workbook("C:\\Users\kimsungju\Desktop\python file\moon_0418.xlsx", data_only=True)
load_wb2 = load_workbook("C:\\Users\kimsungju\Desktop\python file\Ahn_0418.xlsx", data_only=True)    
load_ws1 = load_wb1['sampleSheet']
load_ws2 = load_wb1['sampleSheet']
# To clear the files name 
mpl.rcParams['axes.unicode_minus'] = False
# load the value in the cell 
for i in range(3471):
    num = 'B' + str(i+1)
    col = load_ws1[num].value
    pro_list = []
    for j in gensim.utils.simple_preprocess(col) :
        pro_list.append(j)
    text.append(pro_list)   
for i in range(3362):
    num = 'B' + str(i+1)
    col = load_ws2[num].value
    pro_list = []
    for j in gensim.utils.simple_preprocess(col) :
        pro_list.append(j)
    text.append(pro_list)     
Word = Word2Vec(text, size=10, window=5, min_count=1, workers=2)
Word.train(text, total_examples=len(text), epochs=100)
# 학습이 완료 되면 필요없는 메모리를 unload 시킨다. 
Word.init_sims(replace=True)
# naming the model's name 
model_name = '0413'
Word.save(model_name)
print('Done')
